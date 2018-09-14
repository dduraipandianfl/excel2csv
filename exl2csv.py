import openpyxl
import os.path
import sys

if len(sys.argv) == 1:
    print(">> Syntax error <<")
    print("Usage:: python exl2csv.py file_name(_with_full_path_in_double_quotes)")
else:
    #wb = openpyxl.load_workbook('exceltest.xlsx')
    xlfile = sys.argv[1].rstrip()
    
    if os.path.isfile(xlfile):
        filename, file_extension = os.path.splitext(xlfile)
        
        if file_extension == '.xls' or file_extension == '.xlsx':
            wb = openpyxl.load_workbook(xlfile)
            wbSheets = wb.sheetnames
            print(" >> Sheets in the given excel << ")

            for sheet in wbSheets:
                print("Sheet Title: " + sheet.title())    
                print("Max rows in the sheet:  " + str(wb[sheet].max_row))    
                print("Max columns in the sheet:  " + str(wb[sheet].max_column))
                
                with open(sheet.title() + ".csv",'w') as f:
                    for row in range(1,wb[sheet].max_row+1):
                        rowdata = ""
                        for column in range(1,wb[sheet].max_column+1):
                            rowdata += wb[sheet].cell(row ,column).value + ","
                        f.write(rowdata[:-1] + "\n")
                print("{} sheet is exported to csv. Please check in the current script directory.".format(sheet.title()))
        else:
            print("non-supported file extenstion : " + file_extension)  
    else:
        print("file does not exist")   